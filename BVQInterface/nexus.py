

import datetime as dt
import pandas as pd
import os
from openpyxl import load_workbook

#****************************************************************************************************************
#The names of the business that we have in the historic csv
business= ['BANCO GUAYAQUIL S.A.', 'MUTUALISTA PICHINCHA',
       'CERVECERIA NACIONAL CN S A', 'CORPORACION FAVORITA C.A.',
       'BANCO DE LA PRODUCCION S.A . PRODUBANCO', 'BANCO PICHINCHA C.A.',
       'HOLCIM ECUADOR S.A.', 'SIEMPREVERDE SA', 'TECATEAK SA',
       'CONCLINA C A  CIA CONJU CLINICO NACIONAL', 'RETRATOREC S.A.',
       'BANCO BOLIVARIANO C.A.', 'TECAFORTUNA SA',
       'BOLSA DE VALORES DE GUAYAQUIL', 'SAN CARLOS SOC. AGR. IND.',
       'CRIDESA', 'INDUSTRIAS ALES', 'BOLSA DE VALORES DE QUITO',
       'HOLDING TONICORP S.A.', 'BEVERAGE BRAND PATENTS SA',
       'LA SABANA FORESTAL (PLAINFOREST) S.A.', 'VALLE GRANDE FORESTAL',
       'LA CUMBRE FORESTAL PEAKFOREST SA', 'INVERSANCARLOS',
       'BRIKAPITAL SA', 'MULTI-BG S.A. CORPORACION', 'HOTEL COLON',
       'CERRO ALTO HIGHFOREST S A', 'MERIZA', 'EL TECAL',
       'LA VANGUARDIA FORESTAL', 'NATLUK SA',
       'CONTINENTAL TIRE ANDINA S A',
       'CIALCO CONSTRUCTORA IMPORTADORA ALVAREZ',
       'LA CAMPINA FORESTAL STRONGFOREST S A', 'CEPSA',
       'BANCO SOLIDARIO S.A.', 'BANCO AMAZONAS', 'SUPERDEPORTE S.A.',
       'BANCO DE MACHALA S.A.', 'BANCO DEL AUSTRO',
       'CERRO VERDE FORESTAL S A BIGFOREST',
       'RIO GRANDE FORESTAL RIVERFOREST SA',
       'LA ENSENADA FORESTAL COVEFORESTS SA', 'RIO CONGO FORESTAL',
       'LA COLINA FORESTAL (HILLFOREST) S.A.',
       'LA ESTANCIA FORESTAL FORESTEAD S A', 'VERDETEKA','ALICOSTA BK HOLDING S A', 'EDUCACION Y FUTURO TEKAFUTURO SA']

#Ask for the current date and initialize different list that will help us in the processing process
day_today= str(dt.datetime.today())[:10]
today= dt.date.today()
print(day_today)
oscu= [] #Contains all the paths 
monalisa= [] #Contains all the names of business of the daily scrap in the web
level= {}
today= dt.date.today()
daily_names= ['PRODUBANCO S.A.','INVERSANCARLOS S.A. USD 1,00', 'CORPORACION  FAVORITA C.A', 'BANCO DE GUAYAQUIL USD 1', 'SOC.AG.IND.SAN CARLOS US D  1', 'CONCLINA C.A. ORDINARIA D  1', 'BANCO PICHINCHA CA D 100.00', 'BOLSA DE VALORES DE QUITO', 'PRODUBANCO S.A.', 'CERVECERIA NACIONAL CN S.A.', 'HOLCIM ECUADOR S.A.D 3', 'BEVERAGE BRAND &PATENTS COMP B', 'FONDO INV COTIZADO FIDUCIA', 'BANCO BOLIVARIANO USD 1,00', 'BOLSA DE VALORES DE GUAYAQUIL', 'ASOCIACION MUT.PICHINCHA']
level = {}
file_path = "D:\\DocumentosI\\NGE\\NGE3.0+1.0\\The grid\\acciones180425.xlsx"
sheet_name = "2025"

#****************************************************************************************************************

#Cleans all the extra words in the arrays
def limpiar(nombre):
    irrelevantes = {"S.A.", "USD", "C.A.", "ltda", "S.A", "US","D","1","A","S","DE","S","A","SA"}
    return set([word for word in nombre.split() if word not in irrelevantes])

#Read the current csv file with the names
daily= pd.read_csv(f"D:\\DocumentosI\\NGE\\NGE3.0+1.0\\The grid\\Data\\operaciones-cerradas{day_today}.csv")
#daily= pd.read_csv(f"D:\\DocumentosI\\NGE\\NGE3.0+1.0\\The grid\\Data\\operaciones-cerradas2025-03-28.csv")

#****************************************************************************************************************
#This commented lines help us the first time we scraped all the names of the 'daily' files
    ##yesterday = [f"D:\\DocumentosI\\NGE\\NGE3.0+1.0\\The grid\\Data\\operaciones-cerradas{(today - dt.timedelta(days=i)).strftime("%Y-%m-%d")}.csv" for i in range(90)]
    ##for i in yesterday:
        ##if os.path.exists(i):
            ##oscu.append(i) 


    ##for i in oscu:
        ##provitional= pd.read_csv(i)
        ##for j in provitional["Emisor"].unique():
        ##if j in monalisa:
                ##continue
            ##else:
                ##monalisa.append(j)
#****************************************************************************************************************

def nexxus():
    daily= pd.read_csv(f"D:\\DocumentosI\\NGE\\NGE3.0+1.0\\The grid\\Data\\operaciones-cerradas{day_today}.csv")
    #Verify if there is a new name that is not in the main list
    for j in daily["Emisor"].unique():
        if j in daily_names:
            print("Usual")
            continue
        else:
            daily_names.append(j)
            print("New name")


    #Create two list cleaned
    new_split = [limpiar(i) for i in daily_names]
    old_split = [limpiar(i) for i in business]


    for i, old in enumerate(old_split):
        for j, new in enumerate(new_split):
            interseccion = old & new
            if len(old) <= 3 and len(old) > 1:
                if len(interseccion) >= 2:
                    level[daily_names[j]] = business[i]
            else:
                if len(interseccion) >= 3:
                    level[daily_names[j]] = business[i]

    level['PRODUBANCO S.A.'] = 'BANCO DE LA PRODUCCION S.A . PRODUBANCO'
    level["SOC.AG.IND.SAN CARLOS US D  1"] = 'SAN CARLOS SOC. AGR. IND.'
    level["CONCLINA PREFERIDA 2500 SERIEA"] = 'CONCLINA C A  CIA CONJU CLINICO NACIONAL'
    level["FONDO INV COTIZADO FIDUCIA"]= "FONDO INV COTIZADO FIDUCIA"
    level["CONCLINA PREFERIDASD  1SERIEB"] = 'CONCLINA C A  CIA CONJU CLINICO NACIONAL'
    level['CONCLINA C.A. ORDINARIA D  1']= 'CONCLINA C A  CIA CONJU CLINICO NACIONAL'
    level['INVERSANCARLOS S.A. USD 1,00']= 'INVERSANCARLOS'
    level['ASOCIACION MUT.PICHINCHA']= 'MUTUALISTA PICHINCHA'
    level['FID HOTEL CIUDAD DEL RIO'] = 'FID HOTEL CIUDAD DEL RIO'
    level['FONDO INV COLEC B RAICES UIO 2']= 'FONDO INV COLEC B RAICES UIO'
    level['BRIKAPITAL S.A'] = 'BRIKAPITAL SA'
    level['CORPETROLSA S.A'] = 'CORPETROLSA S.A'

    for i in daily["Emisor"].unique():
        daily.replace({i:level[i]},inplace=True) 

    #******************************************************************************************
    #Create and manipulate columns of the daily dataframe to concat this one to the historic dataframe

    dates= [f"{today.strftime("%d/%m/%Y")}" for i in range(len(daily))]
    zeroes= [0 for i in range(len(daily))]
    daily.drop("Hora Cierre",axis=1,inplace=True)
    daily.insert(0,"FECHA",dates)
    daily.insert(7,"PROCEDENCIA",zeroes)
    daily.columns = ['FECHA', 'EMISOR', 'VALOR','NUMERO ACCIONES', 'VALOR EFECTIVO','PRECIO', 'VALOR NOMINAL' , 'PROCEDENCIA','CASA VENDEDORA','CASA COMPRADORA']
    daily= daily.reindex(columns=['FECHA', 'EMISOR', 'VALOR', 'VALOR NOMINAL', 'PRECIO','NUMERO ACCIONES', 'VALOR EFECTIVO', 'PROCEDENCIA','CASA VENDEDORA','CASA COMPRADORA'])
    daily.insert(0,"Unnamed",zeroes)


    #Append the new data to the main xlsx file
    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        last_row = sheet.max_row
    else:
        last_row = 0  

    with pd.ExcelWriter(
        file_path,
        engine='openpyxl',
        mode='a',
        if_sheet_exists='overlay'
    ) as writer:
        daily.to_excel(writer, sheet_name=sheet_name, startrow=last_row, index=False, header=False)

nexxus()